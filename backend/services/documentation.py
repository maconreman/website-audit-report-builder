"""
Step 6: Documentation — Service layer.
Includes Excel export with conditional formatting.
"""

import os
import pandas as pd
from datetime import datetime

from backend.utils.file_helpers import get_domain_output_folder, get_output_path, read_csv_safe
from backend.session_state import get_session, mark_step_complete


def generate_docs(domain):
    """Generate documentation + Excel export automatically."""
    session = get_session(domain)
    logs = []
    logs.append({"type": "heading", "message": "STEP 6: Generating Documentation"})

    doc = session.get("documentation", {})
    content = _generate_documentation_content(doc)

    folder, clean_domain = get_domain_output_folder(domain)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save text documentation
    txt_filename = f"{clean_domain} - Audit Documentation {timestamp}.txt"
    txt_path = os.path.join(folder, txt_filename)
    with open(txt_path, 'w') as f:
        f.write(content)
    logs.append({"type": "info", "message": f"Documentation saved: {txt_filename}"})

    # Generate Excel export
    xlsx_filename = None
    try:
        xlsx_filename = _generate_excel_export(domain, folder, clean_domain, timestamp)
        if xlsx_filename:
            logs.append({"type": "info", "message": f"Excel export saved: {xlsx_filename}"})
    except Exception as e:
        logs.append({"type": "info", "message": f"Excel export skipped: {e}"})

    mark_step_complete(domain, 6)
    logs.append({"type": "success", "message": "Audit complete"})

    return {
        "status": "complete",
        "filename": txt_filename,
        "filepath": txt_path,
        "xlsx_filename": xlsx_filename,
        "logs": logs,
    }


def _generate_excel_export(domain, folder, clean_domain, timestamp):
    """Generate a formatted .xlsx with conditional formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    audit_path = get_output_path(domain, "audit")
    if not audit_path or not os.path.exists(audit_path):
        return None

    df = read_csv_safe(audit_path)
    wb = Workbook()

    # --- Audit Sheet ---
    ws = wb.active
    ws.title = "Audit"

    # Header style
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Action colors
    action_fills = {
        'Keep': PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        'Remove/Redirect': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'Discuss Further': PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    }

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    action_col_idx = None
    if 'Action' in df.columns:
        action_col_idx = df.columns.tolist().index('Action') + 1

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        action_val = ''
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else '')
            cell.border = thin_border
            cell.font = Font(size=9)
            if col_idx == action_col_idx:
                action_val = str(val) if pd.notna(val) else ''

        # Apply row color based on action
        if action_val in action_fills:
            fill = action_fills[action_val]
            if action_col_idx:
                ws.cell(row=row_idx, column=action_col_idx).fill = fill

    # Auto-width columns (cap at 40)
    for col_idx in range(1, len(df.columns) + 1):
        max_len = len(str(df.columns[col_idx - 1]))
        for row_idx in range(2, min(len(df) + 2, 52)):  # Sample first 50 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Summary Sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Website Audit Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Domain: {clean_domain}")
    ws2.cell(row=3, column=1, value=f"Total Pages: {len(df)}")
    ws2.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    row = 6
    if 'Action' in df.columns:
        ws2.cell(row=row, column=1, value="Action Distribution").font = Font(bold=True)
        row += 1
        for action, count in df['Action'].value_counts().items():
            ws2.cell(row=row, column=1, value=action)
            ws2.cell(row=row, column=2, value=count)
            row += 1
        row += 1

    if 'Page Category' in df.columns:
        ws2.cell(row=row, column=1, value="Page Categories").font = Font(bold=True)
        row += 1
        for cat, count in df['Page Category'].value_counts().items():
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=count)
            row += 1

    xlsx_filename = f"{clean_domain} - Website Audit {timestamp}.xlsx"
    xlsx_path = os.path.join(folder, xlsx_filename)
    wb.save(xlsx_path)

    return xlsx_filename


def get_audit_download_path(domain):
    """Return path to the final Website Audit CSV."""
    path = get_output_path(domain, "audit")
    if path and os.path.exists(path):
        return path
    return None


def get_docs_download_path(domain):
    """Return path to the most recent documentation file."""
    folder, clean_domain = get_domain_output_folder(domain)
    if not folder or not os.path.exists(folder):
        return None

    # Find most recent doc file
    doc_files = [
        f for f in os.listdir(folder)
        if f.startswith(f"{clean_domain} - Audit Documentation") and f.endswith('.txt')
    ]

    if not doc_files:
        return None

    doc_files.sort(reverse=True)
    return os.path.join(folder, doc_files[0])


def _generate_documentation_content(doc):
    """
    Generate documentation content as formatted text.
    Ported 1:1 from generate_documentation_content().
    """
    content = []
    content.append("WEBSITE AUDIT DOCUMENTATION")
    content.append("=" * 60)
    content.append("")
    content.append(f"Domain: {doc.get('domain', 'N/A')}")
    content.append(f"Generated: {doc.get('timestamp', 'N/A')}")
    content.append("")

    # Section 1: Data Sources
    content.append("-" * 60)
    content.append("1. DATA SOURCES PROCESSED")
    content.append("-" * 60)
    content.append("")
    content.append(f"  Screaming Frog (SF-200):")
    content.append(f"    - HTML pages with 200 status: {doc.get('sf_200_rows', 'N/A')}")
    content.append("")

    ga4_info = doc.get('ga4_merge_info', {})
    if ga4_info:
        content.append(f"  GA4 Organic Data:")
        if 'status' in ga4_info:
            content.append(f"    - Status: {ga4_info['status']}")
        else:
            content.append(f"    - Rows in source: {ga4_info.get('rows_merged', 'N/A')}")
            content.append(f"    - Rows matched: {ga4_info.get('rows_matched', 'N/A')}")
            if ga4_info.get('columns_deleted'):
                content.append(f"    - Columns deleted: {', '.join(ga4_info['columns_deleted'])}")
            if ga4_info.get('columns_renamed'):
                for old, new in ga4_info['columns_renamed'].items():
                    content.append(f"    - Renamed: {old} -> {new}")
    content.append("")

    ext_info = doc.get('external_links_merge_info', {})
    if ext_info:
        content.append(f"  External Links Data:")
        if 'status' in ext_info:
            content.append(f"    - Status: {ext_info['status']}")
        else:
            content.append(f"    - Rows in source: {ext_info.get('rows_in_source', 'N/A')}")
            content.append(f"    - Merge key: {ext_info.get('target_column_used', 'N/A')}")
            if ext_info.get('columns_removed'):
                content.append(f"    - Columns removed: {', '.join(ext_info['columns_removed'])}")
    content.append("")

    # Section 2: Custom Data
    content.append("-" * 60)
    content.append("2. CUSTOM DATA COLUMNS")
    content.append("-" * 60)
    content.append("")

    custom_included = doc.get('custom_data_included', [])
    if custom_included:
        for custom_type in custom_included:
            details = doc.get('custom_data_details', {}).get(custom_type, {})
            content.append(f"  {custom_type}:")
            content.append(f"    - Columns combined: {details.get('columns_combined', 1)}")
            if details.get('source_columns'):
                content.append(f"    - Source columns: {', '.join(details['source_columns'])}")
        content.append("")
    else:
        content.append("  No custom data columns included.")
        content.append("")

    # Section 3: Page Categories
    content.append("-" * 60)
    content.append("3. PAGE CATEGORIES ASSIGNED")
    content.append("-" * 60)
    content.append("")

    category_approvals = doc.get('category_approvals', {})
    if category_approvals:
        content.append("  Pattern Approvals:")
        for pattern, category in category_approvals.items():
            pattern_value = pattern.split(':')[1] if ':' in pattern else pattern
            content.append(f"    - {pattern_value} -> {category}")
        content.append("")

    category_summary = doc.get('category_summary', {})
    if category_summary:
        content.append("  Category Summary:")
        for category, count in category_summary.items():
            content.append(f"    - {category}: {count} pages")
        content.append("")

    # Section 4: Action Thresholds
    content.append("-" * 60)
    content.append("4. ACTION RULES AND THRESHOLDS")
    content.append("-" * 60)
    content.append("")

    content.append("  Automatic Rules Applied:")
    content.append("    - All metrics zero -> Remove/Redirect")
    content.append("    - URL contains /tag or /category -> Discuss Further")
    content.append("      (Nexus Note: 'Is this tag/category still relevant?')")
    content.append("    - Has at least 1 lead -> Keep")
    content.append("")

    old_content = doc.get('old_content_settings', {})
    if old_content.get('enabled'):
        content.append("  Old Content Rule:")
        content.append(f"    - Cutoff year: {old_content.get('cutoff_year', 'N/A')}")
        content.append(f"    - Date field used: {old_content.get('date_field', 'N/A')}")
        content.append(f"    - Pages flagged: {old_content.get('pages_flagged', 'N/A')}")
        content.append("")

    thresholds = doc.get('action_thresholds', {})
    if thresholds:
        content.append("  Custom Thresholds:")
        for metric, settings in thresholds.items():
            if settings['type'] == 'percentage':
                content.append(f"    - {metric}: Top {settings['input']}% (>= {settings['actual']:,.2f})")
            else:
                content.append(f"    - {metric}: >= {settings['actual']:,.2f}")
        content.append("")

    recent_override = doc.get('recent_content_override', {})
    if recent_override:
        content.append("  Recent Content Safety Check (last 6 months):")
        if recent_override.get('cutoff_date'):
            content.append(f"    - Date range: {recent_override['cutoff_date']} to present")
        if recent_override.get('enabled'):
            content.append(f"    - Pages overridden to 'Keep': {recent_override.get('pages_overridden', 0)}")
        else:
            content.append(f"    - Skipped (kept original actions for {recent_override.get('pages_skipped', 0)} pages)")
        content.append("    - Note: Tag/category pages excluded from this check")
        content.append("")

    # Section 5: Action Summary
    content.append("-" * 60)
    content.append("5. ACTION SUMMARY")
    content.append("-" * 60)
    content.append("")

    action_summary = doc.get('action_summary', {})
    if action_summary:
        for action, count in action_summary.items():
            content.append(f"  - {action}: {count} pages")
        content.append("")

    # Section 6: Final Output
    content.append("-" * 60)
    content.append("6. FINAL OUTPUT")
    content.append("-" * 60)
    content.append("")
    content.append(f"  File: {doc.get('domain', 'N/A')} - Website Audit.csv")
    content.append(f"  Total rows: {doc.get('final_row_count', 'N/A')}")
    content.append(f"  Total columns: {len(doc.get('final_columns', []))}")
    content.append("")

    final_cols = doc.get('final_columns', [])
    if final_cols:
        content.append("  Columns:")
        for col in final_cols:
            content.append(f"    - {col}")

    content.append("")
    content.append("=" * 60)
    content.append("END OF DOCUMENTATION")
    content.append("=" * 60)

    return '\n'.join(content)
